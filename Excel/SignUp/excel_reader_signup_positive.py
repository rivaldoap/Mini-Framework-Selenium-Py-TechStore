from openpyxl import load_workbook

FILE_PATH = r"Excel/SignUp/SignUp - Positive.xlsx"

def signup_positive():
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["SIGNUP_POSITIVE"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai Kolom E
    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC": row[1],
                "URL": row[2],
                "SIGNUP_USERNAME": row[3],
                "SIGNUP_PASSWORD": row[4],
                }
            )

    return all_data