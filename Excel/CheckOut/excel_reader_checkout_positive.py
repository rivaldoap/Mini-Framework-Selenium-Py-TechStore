from openpyxl import load_workbook

FILE_PATH = r"Excel/CheckOut/CheckOut - Positive.xlsx"

def checkout_positive():
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["CHECKOUT_POSITIVE"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC": row[1],
                "PRODUK_1": row[2],
                "WARNA_PRODUK_1": row[3],
                "PRODUK_2": row[4],
                "WARNA_PRODUK_2": row[5],
                "PRODUK_3": row[6],
                "WARNA_PRODUK_3": row[7],
                "FULLNAME": row[8],
                "COUNTRY": row[9],
                "CITY": row[10],
                "CARD_NUMBER": row[11],
                "MONTH": row[12],
                "YEAR": row[13],
                }
            )

    return all_data